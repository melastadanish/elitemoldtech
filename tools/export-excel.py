#!/usr/bin/env python3
"""
export-excel.py — Export Step 1 Requirement Understanding to Excel (.xlsx)
Clean black text on white background. No colour fills.

Usage:
    python3 tools/export-excel.py storage/2026-06/EMT-LM762-Q001-v1

Requires: pip3 install openpyxl pillow
"""

import sys
import json
import io
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    print("Missing: openpyxl  →  pip3 install openpyxl"); sys.exit(1)

try:
    from PIL import Image as PILImage
except ImportError:
    print("Missing: pillow  →  pip3 install pillow"); sys.exit(1)


# ── Row heights ────────────────────────────────────────────────────────────────
RH_TITLE      = 42    # ELITE MOLD TECH heading
RH_SUBTITLE   = 28    # subtitle / badge row
RH_META       = 24    # quote/date row
RH_DIVIDER    = 5     # thin rule
RH_SPACER     = 12    # breathing space between sections
RH_CARD_HDR   = 30    # card title row (FROM / TO / Job Summary)
RH_CARD_FIELD = 30    # each card field row
RH_SEC_TITLE  = 30    # section title bar (Parts, DFM, etc.)
RH_TABLE_HDR  = 34    # parts table header
RH_TABLE_ROW  = 115   # parts data rows (holds images)
RH_INFO_ROW   = 30    # DFM / missing / assumption rows
RH_FOOTER     = 24

IMG_PX = 130

# ── Column layout ──────────────────────────────────────────────────────────────
COLUMNS = [
    ("No.",                    14),
    ("Part Number",            22),
    ("Name",                   15),
    ("Drawing",                22),
    ("Process",                18),
    ("Dimensions\n(L×W×H mm)", 18),
    ("Material",               18),
    ("Surface Treatment",      32),
    ("Qty",                     6),
]
N = len(COLUMNS)

def col(n): return get_column_letter(n)
def LC():   return col(N)

# ── Borders ────────────────────────────────────────────────────────────────────
def thin(c="000000"):  return Side(style="thin",   color=c)
def med(c="000000"):   return Side(style="medium", color=c)
def none():            return Side(style=None)

LIGHT = "CCCCCC"   # light grey for inner cell borders
DARK  = "000000"   # black for outer card / section borders

def cell_border():
    s = thin(LIGHT)
    return Border(left=s, right=s, top=s, bottom=s)

def card_outer(top="medium", bottom="medium"):
    return Border(left=med(DARK), right=med(DARK),
                  top=Side(style=top, color=DARK),
                  bottom=Side(style=bottom, color=DARK))

def card_inner_label(last=False):
    return Border(left=med(DARK), right=thin(LIGHT),
                  top=none(),
                  bottom=med(DARK) if last else thin(LIGHT))

def card_inner_value(last=False):
    return Border(left=thin(LIGHT), right=med(DARK),
                  top=none(),
                  bottom=med(DARK) if last else thin(LIGHT))


# ── Style helpers ──────────────────────────────────────────────────────────────
def fill(c):  return PatternFill("solid", fgColor=c)
def wfill():  return PatternFill(fill_type=None)   # no fill = white
def fnt(bold=False, sz=10, color="000000", name="Calibri"):
    return Font(bold=bold, size=sz, color=color, name=name)
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def rh(ws, row, h):  ws.row_dimensions[row].height = h

def mc(ws, row, c1, c2):
    """Merge cells and return the top-left cell."""
    ws.merge_cells(f"{col(c1)}{row}:{col(c2)}{row}")
    return ws.cell(row=row, column=c1)

def write(ws, row, c1, c2=None, value="", bold=False, sz=10,
          fg="000000", h="left", v="center", wrap=False,
          border=None, fill_color=None):
    if c2 and c2 != c1:
        cell = mc(ws, row, c1, c2)
    else:
        cell = ws.cell(row=row, column=c1)
    cell.value     = value
    cell.font      = fnt(bold=bold, sz=sz, color=fg)
    cell.alignment = aln(h, v, wrap)
    cell.fill      = fill(fill_color) if fill_color else wfill()
    if border:
        cell.border = border
    return cell


# ── Section title (bold uppercase, black top border) ──────────────────────────
def section_title(ws, row, text):
    cell = write(ws, row, 1, N, value=text.upper(), bold=True, sz=12,
                 border=Border(top=med(DARK), bottom=thin(DARK),
                               left=thin(DARK), right=thin(DARK)))
    rh(ws, row, RH_SEC_TITLE)
    return row + 1


# ── Data ───────────────────────────────────────────────────────────────────────
PARTS = [
    ("20-01.01","LM762X54-20-01.01","Base",          "1",  "Laser Cut + Bend",  "SS 303/304 t=3mm",    "❌ Not specified",                       "1"),
    ("20-01.02","LM762X54-20-01.02","Stand",          "2",  "Laser Cut + Bend",  "SS 303/304 t=3mm",    "❌ Not specified",                       "1"),
    ("20-02.01","LM762X54-20-02.01","Shaft",          "3",  "CNC Turning",       "AISI 4340",           "Black Oxidation MIL-DTL-13924 Cl.4",     "1"),
    ("20-02.02","LM762X54-20-02.02","Flange",         "4",  "CNC Turn+Mill",     "AISI 4340",           "Black Oxidation MIL-DTL-13924 Cl.4",     "1"),
    ("20-02.03","LM762X54-20-02.03","Flange disc",    "5",  "Laser Cut",         "SS 303/304 t=5mm",    "Alkaline Oxidation MIL-DTL-13924 Cl.4",  "1"),
    ("20-02.04","LM762X54-20-02.04","Pin",            "6",  "CNC Turning",       "AISI 4340",           "Black Oxidation MIL-DTL-13924 Cl.4",     "2"),
    ("20-03.01","LM762X54-20-03.01","Locker base",    "7",  "Laser Cut + Bend",  "SS 303/304 t=2mm",    "Alkaline Oxidation MIL-DTL-13924 Cl.4",  "1"),
    ("20-03.02","LM762X54-20-03.02","Locker handle",  "8",  "Laser Cut",         "SS 303/304 t=2mm",    "Alkaline Oxidation MIL-DTL-13924 Cl.4",  "1"),
    ("20-04.01","LM762X54-20-04.01","Spool",          "9",  "CNC Turn+Mill",     "Al 6061",             "Black Anodize",                          "2"),
    ("20-04.02","LM762X54-20-04.02","Wall",           "10", "Laser Cut",         "Textolite/FR4 t=3mm", "Paint Black",                            "4"),
    ("15-07.01","LM762X54-15-07.01","Shaft (sub)",    "11", "CNC Turning",       "AISI 4340",           "Black Oxidation MIL-DTL-13924 Cl.4",     "1"),
    ("15-07.02","LM762X54-15-07.02","Flange (sub)",   "12", "3D Print",          "Polyamide Black",     "As-printed",                             "1"),
    ("15-07.03","LM762X54-15-07.03","Spring (sub)",   "13", "Custom Spring",     "SS wire ø1mm",        "HRC 46-52",                              "2"),
    ("15-08.01","LM762X54-15-08.01","Handle base",    "14", "Laser Cut + Bend",  "SS 303/304 t=2mm",    "Alkaline Oxidation MIL-DTL-13924 Cl.4",  "1"),
    ("15-08.02","LM762X54-15-08.02","Handle knob",    "15", "Laser Cut",         "SS 303/304 t=2mm",    "Alkaline Oxidation MIL-DTL-13924 Cl.4",  "1"),
]

DFM_FLAGS = [
    ("HIGH", "Spring (15-07.03)",         "Custom spring — specialty manufacturing, likely subcontract. Confirm wire OD, free length, coil count, spring rate."),
    ("HIGH", "Wall (20-04.02)",           "Textolite/FR4 laser cutting produces toxic fumes — requires fume extraction or subcontract to specialist."),
    ("MED",  "Flange (15-07.02)",         "3D print bore tolerance ±0.02mm too tight for SLS/MJF — post-machining likely required."),
    ("MED",  "AISI 4340 parts (5 parts)", "High-strength steel — harder to machine, longer cycle times, higher tooling wear."),
    ("LOW",  "All laser-cut SS parts",    "Standard material, straightforward processing. Alkaline oxidation is a common finish."),
]

MISSING_INFO = [
    "Client company name and contact details — needed for the formal quote header",
    "Quantity of complete assemblies — drives all pricing and lead times",
    "Delivery deadline — required for lead time feasibility check",
    "Surface finish for Base (20-01.01) and Stand (20-01.02) — no finish specified",
    "3D print technology for Flange (15-07.02) — SLS or MJF?",
]

ASSUMPTIONS = [
    '"SS 303/304/314" assumed to be SS 304 where unspecified',
    "Quantities based on 1 complete assembly set as shown in BOM",
    "Spring (15-07.03) will be subcontracted — not in-house capability",
    "Textolite Wall (20-04.02) will be subcontracted to PCB/laminate specialist",
]


# ── Image ──────────────────────────────────────────────────────────────────────
def load_xl_image(path, size=IMG_PX):
    try:
        pil = PILImage.open(path).convert("RGB")
        pil.thumbnail((size, size), PILImage.LANCZOS)
        buf = io.BytesIO()
        pil.save(buf, "JPEG", quality=85)
        buf.seek(0)
        img = XLImage(buf)
        img.width = size; img.height = size
        return img
    except Exception as e:
        print(f"    Image error: {e}"); return None


def dim_str(key, dims):
    d = dims.get(key)
    return f"{d['L']} × {d['W']} × {d['H']}" if d else "—"


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Usage: python3 tools/export-excel.py <job-folder>"); sys.exit(1)

    job_folder  = Path(sys.argv[1])
    renders_dir = job_folder / "client-files" / "renders"
    if not renders_dir.exists():
        print(f"Error: {renders_dir} not found — run render-parts.py first"); sys.exit(1)

    manifest = json.loads((renders_dir / "manifest.json").read_text()) \
        if (renders_dir / "manifest.json").exists() else {}
    dimensions = json.loads((renders_dir / "dimensions.json").read_text()) \
        if (renders_dir / "dimensions.json").exists() else {}

    job_name = job_folder.name
    out_path = job_folder / "step1-new.xlsx"

    print(f"\nExporting: {job_name}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Requirement Understanding"

    for i, (_, w) in enumerate(COLUMNS, 1):
        ws.column_dimensions[col(i)].width = w

    row = 1

    # ══════════════════════════════════════════════════════════════════════════
    # HEADER
    # ══════════════════════════════════════════════════════════════════════════

    # Row 1 — Company name (left) | Document title (right)
    write(ws, row, 1, 4, "ELITE MOLD TECH", bold=True, sz=24, h="left")
    write(ws, row, 5, N, "REQUIREMENT UNDERSTANDING", bold=True, sz=18, h="right")
    rh(ws, row, RH_TITLE); row += 1

    # Row 2 — Subtitle (left) | INTERNAL label (right, italic)
    write(ws, row, 1, 4, "Shenzhen Elite Technology Co., Ltd", sz=12, fg="555555")
    write(ws, row, 5, N, "INTERNAL  —  STEP 1 OF 2", sz=12, fg="555555", h="right")
    rh(ws, row, RH_SUBTITLE); row += 1

    # Row 3 — Quote / Date right
    write(ws, row, 5, N, f"Quote: {job_name}    |    Date: 2026-06-18",
          sz=12, fg="555555", h="right")
    rh(ws, row, RH_META); row += 1

    # Row 4 — Black divider
    ws.merge_cells(f"A{row}:{LC()}{row}")
    ws.row_dimensions[row].height = RH_DIVIDER
    for c in range(1, N + 1):
        ws.cell(row=row, column=c).border = Border(top=med(DARK))
    row += 1

    # Spacer
    rh(ws, row, RH_SPACER); row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # INFO CARDS  (cols 1-3 | 4-6 | 7-9)
    # ══════════════════════════════════════════════════════════════════════════

    def card_title(r, c1, c2, label):
        write(ws, r, c1, c2, label, bold=True, sz=13,
              border=Border(left=med(DARK), right=med(DARK),
                            top=med(DARK), bottom=thin(LIGHT)))
        rh(ws, r, RH_CARD_HDR)

    def card_field(r, c1, c2, label, value, last=False):
        # label in first col
        lbl = ws.cell(row=r, column=c1, value=label + ":")
        lbl.font = fnt(sz=12, color="555555"); lbl.fill = wfill()
        lbl.alignment = aln("left", "center")
        lbl.border = card_inner_label(last)

        # value in remaining cols
        ws.merge_cells(f"{col(c1+1)}{r}:{col(c2)}{r}")
        val = ws.cell(row=r, column=c1+1, value=value)
        val.font = fnt(sz=12); val.fill = wfill()
        val.alignment = aln("left", "center", wrap=True)
        val.border = card_inner_value(last)
        rh(ws, r, RH_CARD_FIELD)

    card_title(row, 1, 3, "  FROM (Client)")
    card_title(row, 4, 6, "  TO (Elite Mold Tech)")
    card_title(row, 7, 9, "  Job Summary")
    row += 1

    card_data = [
        [("Company",        "❌ Unknown (OPM-Group?)"),
         ("Company",        "Shenzhen Elite Technology Co., Ltd"),
         ("Assembly",       "LM762X54-20-00.00 — Rounds Loader")],
        [("Contact",        "Armen (designer)"),
         ("Contact",        "Sales / Engineering"),
         ("Files received", "47 (STEP + PDF + DXF + BOM)")],
        [("Email",          "❌ Not provided"),
         ("Quote No",       "EMT-LM762-Q001"),
         ("Custom parts",   "15  (8 standard excluded)")],
        [("Phone",          "❌ Not provided"),
         ("Version",        "v1 — Initial review"),
         ("Quantity",       "❌ Not specified")],
    ]

    for i, trio in enumerate(card_data):
        last = (i == len(card_data) - 1)
        card_field(row, 1, 3, trio[0][0], trio[0][1], last)
        card_field(row, 4, 6, trio[1][0], trio[1][1], last)
        card_field(row, 7, 9, trio[2][0], trio[2][1], last)
        row += 1

    # Spacer
    rh(ws, row, RH_SPACER); row += 1

    # Deadline — full-width plain row
    write(ws, row, 1, N, "Deadline:   ❌  Not specified",
          sz=12, fg="555555",
          border=Border(left=thin(LIGHT), right=thin(LIGHT),
                        top=thin(LIGHT), bottom=thin(LIGHT)))
    rh(ws, row, RH_CARD_FIELD); row += 1

    rh(ws, row, RH_SPACER); row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # PARTS TABLE
    # ══════════════════════════════════════════════════════════════════════════
    row = section_title(ws, row, "Parts Scope — 15 Custom Parts   "
                                  "(Dimensions from STEP bounding box  L × W × H mm)")

    hdr_row = row
    for i, (label, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=hdr_row, column=i, value=label)
        cell.font      = fnt(bold=True, sz=12)
        cell.alignment = aln("center", "center", wrap=True)
        cell.border    = Border(left=thin(LIGHT), right=thin(LIGHT),
                                top=thin(DARK), bottom=med(DARK))
    rh(ws, hdr_row, RH_TABLE_HDR); row += 1

    print(f"  Writing {len(PARTS)} parts...")
    for i, (key, pn, name, no, process, mat, finish, qty) in enumerate(PARTS):
        vals = [no, pn, name, "", process, dim_str(key, dimensions), mat, finish, qty]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font      = fnt(sz=12)
            cell.alignment = aln("center", "center", wrap=True)
            cell.border    = cell_border()

        # Per-column alignment
        ws.cell(row=row, column=1).alignment = aln("center", "center")
        ws.cell(row=row, column=2).alignment = aln("left",   "center")
        ws.cell(row=row, column=3).alignment = aln("left",   "center")
        ws.cell(row=row, column=5).alignment = aln("left",   "center")
        ws.cell(row=row, column=7).alignment = aln("left",   "center")
        ws.cell(row=row, column=8).alignment = aln("left",   "center", wrap=True)
        ws.cell(row=row, column=9).alignment = aln("center", "center")

        ws.cell(row=row, column=1).font = fnt(bold=True, sz=12)
        ws.cell(row=row, column=9).font = fnt(bold=True, sz=12)

        rh(ws, row, RH_TABLE_ROW)

        best = manifest.get(key, {}).get("best")
        if best:
            img = load_xl_image(renders_dir / f"{key}-{best}.jpg")
            if img:
                img.anchor = f"D{row}"; ws.add_image(img)
                print(f"    ✓ {key}")
            else:
                ws.cell(row=row, column=4).value = "No image"
        row += 1

    rh(ws, row, RH_SPACER); row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # DFM FLAGS
    # ══════════════════════════════════════════════════════════════════════════
    row = section_title(ws, row, "DFM Flags")

    for severity, part_lbl, note in DFM_FLAGS:
        # Severity (col 1)
        cell = ws.cell(row=row, column=1, value=severity)
        cell.font = fnt(bold=True, sz=12); cell.alignment = aln("center", "center")
        cell.border = cell_border()

        # Part label (cols 2-3)
        ws.merge_cells(f"B{row}:C{row}")
        pc = ws.cell(row=row, column=2, value=part_lbl)
        pc.font = fnt(bold=True, sz=12); pc.alignment = aln("left", "center")
        pc.border = cell_border()

        # Note (cols 4-9)
        ws.merge_cells(f"D{row}:{LC()}{row}")
        nc = ws.cell(row=row, column=4, value=note)
        nc.font = fnt(sz=12); nc.alignment = aln("left", "center", wrap=True)
        nc.border = cell_border()

        rh(ws, row, RH_INFO_ROW); row += 1

    rh(ws, row, RH_SPACER); row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # MISSING INFORMATION
    # ══════════════════════════════════════════════════════════════════════════
    row = section_title(ws, row, "Missing Information — Required Before Quoting")
    for item in MISSING_INFO:
        write(ws, row, 1, N, f"•   {item}", sz=12, wrap=True,
              border=cell_border())
        rh(ws, row, RH_INFO_ROW); row += 1

    rh(ws, row, RH_SPACER); row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # ASSUMPTIONS
    # ══════════════════════════════════════════════════════════════════════════
    row = section_title(ws, row, "Assumptions Made")
    for item in ASSUMPTIONS:
        write(ws, row, 1, N, f"•   {item}", sz=12, wrap=True,
              border=cell_border())
        rh(ws, row, RH_INFO_ROW); row += 1

    rh(ws, row, RH_SPACER); row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # FOOTER
    # ══════════════════════════════════════════════════════════════════════════
    write(ws, row, 1, N,
          f"{job_name}  |  Generated: 2026-06-18  |  INTERNAL USE ONLY  —  Elite Mold Tech",
          sz=11, fg="888888", h="center",
          border=Border(top=thin(DARK), bottom=thin(DARK),
                        left=thin(LIGHT), right=thin(LIGHT)))
    rh(ws, row, RH_FOOTER)

    # ── Freeze & print ────────────────────────────────────────────────────────
    ws.freeze_panes = f"A{hdr_row + 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"${hdr_row}:${hdr_row}"

    wb.save(str(out_path))
    sz = out_path.stat().st_size // 1024
    print(f"\n✅ Saved: {out_path}  ({sz} KB)")


if __name__ == "__main__":
    main()
